from datetime import date
import os.path
from typing import Any
import pandas as pd
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import functions
from app_logging import logger


HEADERS = [
    "Kuupäev",
    "kinnisvara puhas väärtus",
    "Füüsilise isiku aktsiad",
    "Juriidilise isiku aktsiad",
    "aktsiad kokku",
    "Terve portfell kokku",
    "Mörr-i portfell",
    "Pere portfell kokku",
    "Vilde after Tax",
    "Vaba Raha",
    "Funderbeam Väärtus",
    "Kelly portfell kokku",
]


def get_excel_path(excel_name: str) -> str:
    base_path = functions.what_path_for_file()
    if not base_path:
        # Use the directory of this script as fallback
        base_path = os.path.abspath(os.path.dirname(__file__))
    file_name = excel_name + ".xlsx"
    file_path = os.path.join(base_path, "data", file_name)
    file_path = os.path.abspath(file_path)
    logger.debug(f"Excel file path: {file_path}")
    return file_path


def freeze_excel_row(excel_name: str) -> None:
    file_path = get_excel_path(excel_name)
    wb = load_workbook(file_path)
    sheet1 = wb.active
    sheet1.freeze_panes = "A2"  # freeze first row
    wb.save(filename=file_path)


def create_excel(excel_name: str, sheet_name: str) -> None:
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_name
    file_path = get_excel_path(excel_name)
    wb.save(filename=file_path)
    wb.close()
    print("=================================================================================")
    print("Loodud uus fail", file_path)


def check_if_excel_exists(excel_name: str) -> bool:
    return bool(get_excel_path(excel_name))


# how_to_add: 1 = append, 2 = overwrite,
# compare_column for overwrite: 1 is first column in excel (A) and 2 is B and so on
def write_to_excel(excel_name: str, list_of_data: list, how_to_add: int = None, compare_column: int = None) -> None:
    file_path = get_excel_path(excel_name)
    wb = load_workbook(file_path)
    sheet1 = wb.active
    # just append a row'
    if how_to_add == 1:
        sheet1.append(list_of_data)
        print("Tänane seis lisatud.")
    elif how_to_add == 2:
        max_row = sheet1.max_row
        # Overwrite row if compared column value (ex. date) is the same as given data column'
        if sheet1.cell(column=compare_column, row=max_row).value == list_of_data[compare_column - 1]:
            # easier to delete row and append looks easier than replace cell by cell'
            sheet1.delete_rows(max_row)
            sheet1.append(list_of_data)
            print("Tänane seis üle kirjutatud")
        else:
            sheet1.append(list_of_data)
            print("Tänane seis lisatud.")
    wb.save(filename=file_path)
    wb.close()


def column_width(excel_name: str, excel_headers: list) -> None:
    file_path = get_excel_path(excel_name)
    wb = load_workbook(file_path)
    sheet1 = wb.active
    for i, col_value in enumerate(excel_headers, 1):
        # if column length is very small (less then 5), then give static length of 10, else length of column'
        if len(col_value) < 5:
            column_extender = 10
        else:
            column_extender = len(col_value)
        # wants column letter for input, as i. Width input is in the end of it'
        sheet1.column_dimensions[get_column_letter(i)].width = column_extender
    wb.save(filename=file_path)
    wb.close()


# check if excel file is there, if not create it
def need_new_excel_file(excel_name: str, sheet_name: str, excel_headers: list) -> None:
    if check_if_excel_exists(excel_name):
        print("=================================================================================")
        print("Fail juba kaustas olemas.")
    else:
        create_excel(excel_name, sheet_name)
        freeze_excel_row(excel_name)
        write_to_excel(excel_name, HEADERS)
        column_width(excel_name, excel_headers)


def get_excel_column_values(excel_name: str, column_letter: str) -> list:
    file_path = get_excel_path(excel_name)
    wb = load_workbook(file_path)
    sheet1 = wb.active
    column_list = []
    # using enumerate to get index and then to skip header row
    for index, col in enumerate(sheet1[column_letter]):
        if index == 0:
            continue
        column_list.append(col.value)

    return column_list


def get_last_row(excel_name: str, column_number: int) -> Any:
    logger.debug(f"excel name is {excel_name}, column number is {column_number}")
    file_path = get_excel_path(excel_name)
    wb = load_workbook(file_path)
    sheet1 = wb.active

    max_rows = sheet1.max_row
    cell = sheet1.cell(row=max_rows, column=column_number).value
    logger.debug(f"Checked max_row={max_rows}, cell value={cell}")

    if cell not in (None, ''):
        return cell

    # If the last row is empty, scan upwards for the nearest non-empty cell
    # Issue is that if excel file has been modified manually, then last row might be empty for example onlyOffice etc
    for row in range(max_rows - 1, 0, -1):
        cell = sheet1.cell(row=row, column=column_number).value
        if cell not in (None, ''):
            logger.debug(f"Found last non-empty cell at row {row} with value {cell}")
            return cell

    logger.debug("No non-empty cells found in the column")
    return None


def year_to_year_percent(
    excel_name: str, mm_dd: str, todays_total_portfolio: float, portfolio_history_column: str, filter_nr_input: int = 0
) -> pd.DataFrame:
    # Load the workbook and get the data
    dates = get_excel_column_values(excel_name, "A")
    amounts = get_excel_column_values(excel_name, portfolio_history_column)

    # Create a DataFrame from the dates and amounts
    df = pd.DataFrame({"Date": dates, "Amount": amounts})

    if df.empty:
        logger.error("Dataframe is empty for year to year percent calculation")
        return pd.DataFrame()

    # Filter the DataFrame based on the mm_dd condition
    df = df[df["Date"].str.contains(mm_dd, na=False)]

    # Add today's portfolio amount if conditions are met
    today = date.today()
    last_date = parse(df["Date"].iloc[-1]).date()  # Get the last date in the DataFrame

    # Always append today's portfolio amount if the year matches
    if today.year == last_date.year:
        new_row = pd.DataFrame({"Date": [today], "Amount": [round(todays_total_portfolio)]})
        df = pd.concat([df, new_row], ignore_index=True)

    # Calculate previous amounts and percentage increases
    df["Previous Amount"] = df["Amount"].shift(1)
    df["Percentage Increase"] = ((df["Amount"] - df["Previous Amount"]) / df["Previous Amount"] * 100).fillna(0).round()

    # final ouput and onvert to integer to remove decimal places
    data = {
        "Aasta": df["Date"],
        "Portfell eelmisel aastal": df["Previous Amount"].fillna(0).astype(int),
        "Portfell see aasta": df["Amount"].astype(int),
        "Protsendiline muutus": df["Percentage Increase"].astype(int).astype(str) + " %",
    }

    final_df = pd.DataFrame(data)

    pd.set_option("display.max_rows", None)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", None)

    # Filter the DataFrame based on the specified condition
    final_df = final_df[final_df["Portfell see aasta"] >= filter_nr_input].reset_index(drop=True)

    # Replace the last "Aasta" column value with 'Täna'
    if not final_df.empty:
        final_df.iloc[-1, final_df.columns.get_loc("Aasta")] = "Täna"

    return final_df
