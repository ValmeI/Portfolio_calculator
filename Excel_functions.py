import os.path
from datetime import date

import pandas as pd
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import Functions

HEADERS = ["Kuupäev",
           "Kinnisvara puhas väärtus",
           "Füüsilise isiku aktsiad",
           "Juriidilise isiku aktsiad",
           "Aktsiad kokku",
           "Terve portfell kokku",
           "Mörr-i portfell",
           "Pere portfell kokku",
           "Vilde after Tax",
           "Vaba Raha",
           "Funderbeam Väärtus",
           "Kelly portfell kokku"]


def freeze_excel_row(excel_name):
    file_name = excel_name + ".xlsx"
    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active
    sheet1.freeze_panes = 'A2' # freeze first row
    wb.save(filename=workbook_name)


def create_excel(excel_name, sheet_name):
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_name
    file_name = excel_name + ".xlsx"
    wb.save(filename=file_name)
    wb.close()
    print("==================================================")
    print("Loodud uus fail", file_name)


def check_if_excel_exists(excel_name):
    if os.path.isfile(Functions.what_path_for_file() + 'Portfolio_calculator/' + excel_name + ".xlsx"):
        return True
    else:
        return False


# append new rows/info to excel
# how_to_add: 1 = append, 2 = overwrite, 3 = compare if change is needed
# compare_column for overwrite: 1 is first column in excel (A) and 2 is B and so on
def write_to_excel(excel_name, list_of_data, how_to_add, compare_column):
    file_name = excel_name + ".xlsx"
    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active
    # just append a row'
    if how_to_add == 1:
        sheet1.append(list_of_data)
        print("Tänane seis lisatud.")
    elif how_to_add == 2:
        max_row = sheet1.max_row
        # Overwrite row if compared column value (ex. date) is the same as given data column'
        if sheet1.cell(column=compare_column, row=max_row).value == list_of_data[compare_column-1]:
            # easier to delete row and append looks easier than replace cell by cell'
            sheet1.delete_rows(max_row)
            sheet1.append(list_of_data)
            print("Tänane seis üle kirjutatud")
        else:
            sheet1.append(list_of_data)
            print("Tänane seis lisatud.")

    # TODO compare cell by cell if any change is acutally needed
    elif how_to_add == 3:
        print('compare')

    wb.save(filename=workbook_name)


def column_width(excel_name, excel_headers):
    file_name = excel_name + ".xlsx"
    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active
    for i, col_value in enumerate(excel_headers, 1):
        # if column length is very small (less then 5), then give static length of 10, else length of column'
        if len(col_value) < 5:
            column_extender = 10
        else:
            column_extender = len(col_value)
        # wants column letter for input, as i. Width input is in the end of it'
        sheet1.column_dimensions[get_column_letter(i)].width = column_extender
    wb.save(filename=workbook_name)


# check if excel file is there, if not create it
def need_new_excel_file(excel_name, sheet_name, excel_headers):
    if check_if_excel_exists(excel_name):
        print("==================================================")
        print("Fail juba kaustas olemas.")
    else:
        create_excel(excel_name, sheet_name)
        freeze_excel_row(excel_name)
        write_to_excel(excel_name, HEADERS)
        column_width(excel_name, excel_headers)


# returns all values of given column in list
def get_excel_column_values(excel_name, column_letter):
    # add file type
    file_name = excel_name + ".xlsx"
    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active
    column_list = []
    # using enumerate to get index and then to skip header row
    for index, col in enumerate(sheet1[column_letter]):
        if index == 0:
            continue
        column_list.append(col.value)

    return column_list


# returns last row of given columns number
def get_last_row(excel_name, column_number):
    # add file type'
    file_name = excel_name + ".xlsx"

    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active

    max_rows = sheet1.max_row
    cell = sheet1.cell(column=column_number, row=max_rows).value
    return cell


def year_to_year_percent(excel_name, mm_dd, todays_total_portfolio, excel_column_input, filter_nr_input=0):
    file_name = excel_name + ".xlsx"
    workbook_name = file_name
    load_workbook(workbook_name).active
    # all dates and all values from total sum of portfolio'
    date_and_sum_dict = dict(zip(get_excel_column_values(excel_name, 'A'), get_excel_column_values(excel_name, excel_column_input)))
    amount_list = []
    date_list = []
    # to filter out only give dates (mm_dd input) and sums'
    for date1, amount in date_and_sum_dict.items():
        if mm_dd in date1:
            amount_list.append(round(amount))
            date_list.append(date1)
            # is same year as last row (for example 2022-01-01) and it is not January 1st, then add today s portfolio amount'
            if date.today().year == parse(date1).date().year and date.today().month != '1' and date.today().day != '1':
                amount_list.append(round(todays_total_portfolio))
                date_list.append(date.today())

    previous_amount_list = []
    percentage_increase_list = []
    # to get previous vs current values and percentage increase'
    for previous, current in zip(amount_list, amount_list[1:]):
        percentage_increase = round(100*((current-previous)/previous))
        previous_amount_list.append(previous)
        percentage_increase_list.append(str(percentage_increase) + ' %')

    # need to add 0 to the beginning of list, so dataframe would have exactly same amount of rows'
    if len(previous_amount_list) != len(date_list):
        # pos and value added'
        previous_amount_list.insert(0, 0)

    if len(percentage_increase_list) != len(date_list):
        # pos and value added'
        percentage_increase_list.insert(0, '0 %')

    data = {"Aasta": date_list,
            "Portfell eelmisel aastal": previous_amount_list,
            "Portfell see aasta": amount_list,
            "Protsendiline muutus": percentage_increase_list}

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    df = pd.DataFrame(data)
    df = df[df['Portfell see aasta'] >= filter_nr_input]
    # replace last Aasta columns value with 'Täna'
    df.iloc[-1, df.columns.get_loc('Aasta')] = 'Täna'
    df.reset_index(drop=True, inplace=True)
    return df
